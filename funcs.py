import time
import cv2
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.styles import Alignment
from scipy.integrate import quad
from scipy.interpolate import CubicSpline


def extract_curve(image_path):
    image = cv2.imread(image_path)
    hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    red_border_low_1 = np.array([0, 50, 20])
    red_border_high_1 = np.array([5, 255, 255])
    red_border_low_2 = np.array([170, 25, 20])
    red_border_high_2 = np.array([180, 255, 255])
    red_mask_1 = cv2.inRange(hsv_image, red_border_low_1, red_border_high_1)
    red_mask_2 = cv2.inRange(hsv_image, red_border_low_2, red_border_high_2)
    red_mask_result = cv2.bitwise_or(red_mask_1, red_mask_2)
    red_only_image = cv2.bitwise_and(image, image, mask=red_mask_result)
    # plt.imshow(red_only_image)
    # plt.show()

    green_border_low = np.array([35, 25, 25])
    green_border_high = np.array([85, 255, 255])
    green_mask = cv2.inRange(hsv_image, green_border_low, green_border_high)
    green_only_image = cv2.bitwise_and(image, image, mask=green_mask)
    # plt.imshow(green_only_image)
    # plt.show()

    red_gray_image = cv2.cvtColor(red_only_image, cv2.COLOR_BGR2GRAY)
    _, red_thresh_image = cv2.threshold(red_gray_image, 30, 255, cv2.THRESH_BINARY)
    red_contours, _ = cv2.findContours(red_thresh_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    # plt.imshow(red_gray_image)
    # plt.show()

    green_gray_image = cv2.cvtColor(green_only_image, cv2.COLOR_BGR2GRAY)
    _, green_thresh_image = cv2.threshold(green_gray_image, 30, 255, cv2.THRESH_BINARY)
    green_contours, _ = cv2.findContours(green_thresh_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    plt.imshow(green_gray_image)
    plt.show()

    return red_contours, green_contours, image


def pixel_to_real_coordinates(pixel_coords, x_range, y_range, img_shape):
    height, width = img_shape[:2]
    real_coords = []
    # sorted_real_coords = []

    x_min, x_max = x_range
    y_min, y_max = y_range

    for point in pixel_coords:
        x_pixel, y_pixel = point[0]

        x_real = x_min + (x_pixel / width) * (x_max - x_min)
        y_real = y_min + (y_pixel / height) * (y_max - y_min)

        real_coords.append([x_real, -y_real])
    return np.array(real_coords)


def evaluate_curve(image_path, x_range, y_range):
    red_contours, green_contours, image = extract_curve(image_path)

    if red_contours:
        red_contour = max(red_contours, key=cv2.contourArea)
        green_contour = max(green_contours, key=cv2.contourArea)

        red_real_coords = pixel_to_real_coordinates(red_contour, x_range, y_range, image.shape)
        green_real_coords = pixel_to_real_coordinates(green_contour, x_range, y_range, image.shape)

        red_x_values = red_real_coords[:, 0]
        red_y_values = red_real_coords[:, 1]

        green_x_values = green_real_coords[:, 0]
        green_y_values = green_real_coords[:, 1]

        red_x_min, red_x_max = red_x_values.min(), red_x_values.max()
        red_y_min, red_y_max = red_y_values.min(), red_y_values.max()

        return red_real_coords, green_real_coords
    else:
        print("No curve found!")


def average_curve(coords):
    coord_dict = {}
    new_coords = []

    for x, y in coords:
        if x not in coord_dict:
            coord_dict[x] = []
        coord_dict[x].append(y)

    for x, y_values in coord_dict.items():
        avg_y = sum(y_values) / len(y_values)
        new_coords.append([x, avg_y])

    return new_coords


def evaluate_square(red_curve_coords, green_curve_coords):
    red_x_vals = [x for x, y in red_curve_coords]
    red_y_vals = [y for x, y in red_curve_coords]
    red_spline_func = CubicSpline(red_x_vals, red_y_vals)

    green_x_vals = [x for x, y in green_curve_coords]
    green_y_vals = [y for x, y in green_curve_coords]
    green_spline_func = CubicSpline(green_x_vals, green_y_vals)

    x_min = max(red_x_vals[0], green_x_vals[0])
    x_max = min(red_x_vals[-1], green_x_vals[-1])

    # Integrate to find the area under the curves
    red_curve_square, _ = quad(red_spline_func, x_min, x_max)
    green_curve_square, _ = quad(green_spline_func, x_min, x_max)

    curve_square = abs(red_curve_square - green_curve_square)

    return curve_square


def div_coords(coords, max_y):
    return [(x, (y - max_y) / max_y) for x, y in coords]


def recalculate_h(coords):
    return [(1 / x, y) for x, y in coords]


def find_closest_point(coords, target_x=None, target_y=None):
    if target_x is not None:
        return min(coords, key=lambda coord: abs(coord[0] - target_x))  # Closest to x=0
    elif target_y is not None:
        return min(coords, key=lambda coord: abs(coord[1] - target_y))  # Closest to y=0
    return None


def save_to_excel(red_data, green_data, div_data, inv_h_data, other_data,
                  img_name):
    df_red = pd.DataFrame({
        'H': [x for x, y in red_data],
        'M': [y for x, y in red_data],
        'δM': [delta_y for x, delta_y in div_data],
        '1/H': [inv_h for inv_h, y in inv_h_data]
    })

    df_green = pd.DataFrame({
        'H': [x for x, y in green_data],
        'M': [y for x, y in green_data]
    })

    time_string = time.strftime("%d.%m.%Y-%H.%M.%S")
    file_path = "data/hysteresis-loop-" + f"{img_name}-" + f"{time_string}" + ".xlsx"

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_combined = pd.concat([df_red, df_green], axis=1)
        df_combined.to_excel(writer, sheet_name='Data', startrow=2, index=False)

        workbook = writer.book
        worksheet = writer.sheets['Data']

        worksheet.merge_cells('A1:D1')
        worksheet['A1'] = 'Main'
        worksheet['A1'].alignment = Alignment(horizontal='center')

        worksheet.merge_cells('E1:F1')
        worksheet['E1'] = 'Secondary'
        worksheet['E1'].alignment = Alignment(horizontal='center')

        params_ws = workbook.create_sheet(title="Parameters")
        params_ws.append(["Parameter", "Value"])

        # Параметры, такие как Ms, Hc
        param_names = ['Ms', 'red_Mr', 'red_Hc', 'green_Mr', 'green_Hc', 'S']
        for name, value in zip(param_names, other_data):
            params_ws.append([name, value])

    print(f"Data saved into: {file_path}")

# def analyse_graph():
#     image_path = 'graphs/loop_3.jpg'
#     image = cv2.imread(image_path)
#     image_name = os.path.basename(image_path)
#     x_range = (-40, 40)
#     y_range = (-80, 80)
#
#     red_coords, green_coords = evaluate_curve(image_path, x_range, y_range)
#     red_sorted_coords = sorted(np.asarray(red_coords).tolist(), key=lambda coord: coord[0])
#     green_sorted_coords = sorted(np.asarray(green_coords).tolist(), key=lambda coord: coord[0])
#     red_average_coords = average_curve(red_sorted_coords)
#     green_average_coords = average_curve(green_sorted_coords)
#     print("red_average_coords")
#     print(red_sorted_coords)
#     print("green_average_coords")
#     print(green_average_coords)
#
#     fig, ax = plt.subplots(1, 3, figsize=(12, 6))
#
#     ax[0].imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
#     ax[0].set_title('Source graph:')
#     ax[0].axis('off')
#     ax[1].plot(red_average_coords[0], red_average_coords[1], lw=2)
#     ax[1].set_title('Rebuilt graph (main)')
#     ax[1].set_xlabel('H, kOe')
#     ax[1].set_ylabel('M, emu/g')
#     ax[1].grid(True)
#     ax[2].plot(green_average_coords[0], green_average_coords[1], lw=2)
#     ax[2].set_title('Rebuilt graph (secondary)')
#     ax[2].set_xlabel('H, kOe')
#     ax[2].set_ylabel('M, emu/g')
#     ax[2].grid(True)
#     plt.tight_layout()
#     plt.show()
#
#     print(f"PARAMS:")
#
#     params = []
#     Ms = red_average_coords[-1][1]
#     Msmin = red_average_coords[0][1]
#
#     # max_y = max(coords, key=lambda coord: coord[1])[1]
#     # min_y = min(coords, key=lambda coord: coord[1])[1]
#
#     Mr_point = find_closest_point(red_coords, target_x=0)
#     Mr = Mr_point[1]
#
#     Hc_point = find_closest_point(red_coords, target_y=0)
#     Hc = Hc_point[0]
#
#     curve_square = evaluate_square(red_average_coords, green_average_coords)
#
#     params.append(Ms)
#     params.append(abs(Mr))
#     params.append(abs(Hc))
#     params.append(curve_square)
#
#     print(f"Ms: {Ms}")
#     print(f"Msmin: {Msmin}")
#     if Mr:
#         print(f"Mr: {Mr}")
#     if Hc:
#         print(f"Hc: {Hc}")
#     print(f"Square:{curve_square}")
#
#     d_coords = div_coords(red_average_coords, Ms)
#     x = [i[0] for i in d_coords]
#     y = [i[1] for i in d_coords]
#     plt.plot(x, y, lw=2)
#     plt.title('Rebuilt graph (secondary)')
#     plt.xlabel('H, kOe')
#     plt.ylabel('M, emu/g')
#     plt.grid(True)
#     plt.show()
#
#     inv_h_coords = recalculate_h(red_average_coords)
#     x = [i[0] for i in inv_h_coords]
#     y = [i[1] for i in inv_h_coords]
#     plt.plot(x, y, lw=2)
#     plt.title('Rebuilt graph (secondary)')
#     plt.xlabel('H, kOe')
#     plt.ylabel('M, emu/g')
#     plt.grid(True)
#     plt.show()
#
#     if red_average_coords and green_average_coords is not None:
#         save_to_excel(red_average_coords, green_average_coords, d_coords, inv_h_coords, params, image_name)
